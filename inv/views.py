from django.shortcuts import render,redirect
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.views import generic
from django.urls import reverse_lazy
from django.views import View
from bases.views import sin_privilegios


from django.contrib.auth.decorators import login_required,permission_required

from .models import Categoria,SubCategoria,Marca,UnidadMedida,Producto,historiareporte
from .forms import CategoriaForm,SubCategoriaForm,MarcaForm,UMForm,ProductoForm,precisionForm
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Side,PatternFill,Font
from django.http.response import HttpResponse

# Create your views here.

# creacion de los view de la categoria
class CategoriaView(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_categoria"
    model=Categoria
    template_name="inv/categoria_list.html"
    context_object_name="obj"
    login_url="bases:login"
class CategoriaNew(LoginRequiredMixin,sin_privilegios,generic.CreateView):
    permission_required="inv.view_categoria"
    model=Categoria
    template_name="inv/categoria_form.html"
    context_object_name="obj"
    form_class=CategoriaForm
    success_url=reverse_lazy("inv:categoria_list")
    login_url="bases:login"
    def form_valid(self,form):
        form.instance.uc=self.request.user
        return super().form_valid(form)

class CategoriaEdit(LoginRequiredMixin,sin_privilegios,generic.UpdateView,):
    permission_required="inv.view_categoria"
    model=Categoria
    template_name="inv/categoria_form.html"
    context_object_name="obj"   
    form_class=CategoriaForm
    success_url=reverse_lazy("inv:categoria_list")
    login_url="bases:login" 
    
    def form_valid(self,form):
        form.instance.um=self.request.user.id
        return super().form_valid(form)

# class CategoriaDel(LoginRequiredMixin,generic.DeleteView):
#     model=Categoria
#     template_name='inv/categoria_del.html'
#     context_object_name='obj'
#     success_url=reverse_lazy("inv:categoria_list")

@login_required(login_url='/login/')
@permission_required('inv.change_catalogo',login_url='bases:sin_privilegios')

def categoria_inactivar(request,id):
    cat=Categoria.objects.filter(pk=id).first()
    contexto={}
    template_name="inv/catalogo_del.html"

    if not cat:
        return redirect("inv:categoria_list")
    if request.method=='GET':
        contexto={'obj':cat}
    if request.method=='POST':
        cat.estado=False
        cat.save()
        return redirect("inv:categoria_list")
    return render(request,template_name,contexto)




# creacion de las view de la subCategoria
class SubCategoriaView(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_subcategoria"
    model=SubCategoria
    template_name="inv/subcategoria_list.html"
    context_object_name="obj"
    login_url="bases:login"
class SubCategoriaNew(LoginRequiredMixin,sin_privilegios,generic.CreateView):
    permission_required="inv.view_subcategoria"
    model=SubCategoria
    template_name="inv/subcategoria_form.html"
    context_object_name="obj"
    form_class=SubCategoriaForm
    success_url=reverse_lazy("inv:subcategoria_list")
    login_url="bases:login"
    def form_valid(self,form):
        form.instance.uc=self.request.user
        return super().form_valid(form)
class SubCategoriaEdit(LoginRequiredMixin,sin_privilegios,generic.UpdateView):
    permission_required="inv.view_subcategoria"
    model=SubCategoria
    template_name="inv/subcategoria_form.html"
    context_object_name="obj"   
    form_class=SubCategoriaForm
    success_url=reverse_lazy("inv:subcategoria_list")
    login_url="bases:login" 
    def form_valid(self,form):
        form.instance.um=self.request.user.id
        return super().form_valid(form)
# class SubCategoriaDel(LoginRequiredMixin,generic.DeleteView):
#     model=SubCategoria
#     template_name='inv/subcategoria_del.html'
#     context_object_name='obj'
#     success_url=reverse_lazy("inv:subcategoria_list")
@login_required(login_url='/login/')
@permission_required('inv.change_subcategoria',login_url='bases:sin_privilegios')

def sub_inactivar(request,id):
    sub=SubCategoria.objects.filter(pk=id).first()
    contexto={}
    template_name="inv/catalogo_del.html"

    if not sub:
        return redirect("inv:subcategoria_list")
    if request.method=='GET':
        contexto={'obj':sub}
    if request.method=='POST':
        sub.estado=False
        sub.save()
        return redirect("inv:subcategoria_list")
    return render(request,template_name,contexto)




# creacion de los view de la marca

class MarcaView(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_marca"
    model=Marca
    template_name="inv/marca_list.html"
    context_object_name="obj"
    login_url="bases:login"
    permission_required="inv.view_marca"

class MarcaNew(LoginRequiredMixin,sin_privilegios,generic.CreateView):
    permission_required="inv.view_marca"
    model=Marca
    template_name="inv/marca_form.html"
    context_object_name="obj"
    form_class=MarcaForm
    success_url=reverse_lazy("inv:marca_list")
    login_url="bases:login"

    def form_valid(self,form):
        form.instance.uc=self.request.user
        return super().form_valid(form)

class MarcaEdit(LoginRequiredMixin,sin_privilegios,generic.UpdateView):
    permission_required="inv.view_marca"
    model=Marca
    template_name="inv/marca_form.html"
    context_object_name="obj"   
    form_class=MarcaForm
    success_url=reverse_lazy("inv:marca_list")
    login_url="bases:login" 
    def form_valid(self,form):
        form.instance.um=self.request.user.id
        return super().form_valid(form)
@login_required(login_url='/login/')
@permission_required('inv.change_marca',login_url='bases:sin_privilegios')
def marca_inactivar(request,id):
    marca =Marca.objects.filter(pk=id).first()
    contexto={}
    template_name="inv/catalogo_del.html"


    if not marca:
        return redirect("inv:marca_list")
    if request.method=='GET':
        contexto={'obj':marca}
    if request.method=='POST':
        marca.estado=False
        marca.save()
        return redirect("inv:marca_list")


    return render(request,template_name,contexto)


#----------------------- creacion de las unidades de medida--------------




class UMView(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_unidadmedida"
    model=UnidadMedida
    template_name="inv/um_list.html"
    context_object_name="obj"
    login_url="bases:login"

class UMNew(LoginRequiredMixin,sin_privilegios,generic.CreateView):
    permission_required="inv.view_unidadmedida"
    model=UnidadMedida
    template_name="inv/um_form.html"
    context_object_name="obj"
    form_class=UMForm
    success_url=reverse_lazy("inv:um_list")
    login_url="bases:login"

    def form_valid(self,form):
        form.instance.uc=self.request.user
        return super().form_valid(form)

class UMEdit(LoginRequiredMixin,sin_privilegios,generic.UpdateView):
    permission_required="inv.view_unidadmedida"
    model=UnidadMedida
    template_name="inv/um_form.html"
    context_object_name="obj"   
    form_class=UMForm
    success_url=reverse_lazy("inv:um_list")
    login_url="bases:login" 
    def form_valid(self,form):
        form.instance.um=self.request.user.id
        return super().form_valid(form)

def um_inactivar(request,id):
    um =UnidadMedida.objects.filter(pk=id).first()
    contexto={}
    template_name="inv/catalogo_del.html"


    if not um:
        return redirect("inv:um_list")
    if request.method=='GET':
        contexto={'obj':um}
    if request.method=='POST':
        um.estado=False
        um.save()
        return redirect("inv:um_list")


    return render(request,template_name,contexto)

    #--------------creacion delos view de productos
class ProductoView(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_producto"
    model=Producto
    template_name="inv/producto_list.html"
    context_object_name='obj'
    login_url="bases:login"

class ProductoNew(LoginRequiredMixin,sin_privilegios,generic.CreateView):
    permission_required="inv.view_producto"
    model=Producto
    template_name="inv/producto_form.html"
    context_object_name="obj"
    form_class=ProductoForm
    success_url=reverse_lazy("inv:producto_list")
    login_url="bases:login"

    def form_valid(self,form):
        form.instance.uc=self.request.user
        return super().form_valid(form)
class ProductoEdit(LoginRequiredMixin,sin_privilegios,generic.UpdateView):
    permission_required="inv.view_producto"
    model=Producto
    template_name="inv/producto_form.html"
    context_object_name='obj'
    form_class=ProductoForm
    success_url=reverse_lazy("inv:producto_list")
    login_url="bases:login"
    
    def form_valid(self,form):
        form.instance.um=self.request.user.id
        return super().form_valid(form)


def producto_inactivar(request,id):
    prod=Producto.objects.filter(pk=id).first()
    contexto={}
    template_name="inv/catalogo_del.html"

    if not prod:
        return redirect("inv:producto_list")
    if request.method=='GET':
        contexto={'obj':prod}
    if request.method=='POST':
        prod.estado=False
        prod.save()
        return redirect("inv:producto_list")
    return render(request,template_name,contexto)



class reporteprecisionView(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_producto"
    model=Producto
    template_name="inv/reporte_preci.html"
    context_object_name="obj"
    login_url="bases:login"

class guardarprecision(LoginRequiredMixin,sin_privilegios,generic.CreateView):
    permission_required="inv.view_categoria"
    model=Producto
    template_name="inv/guardar_precision.html"
    context_object_name="obj"
    login_url="bases:login"
    form_class=precisionForm
    success_url=reverse_lazy("inv:precision_lista")
    login_url="bases:login"
    def form_valid(self,form):
        form.instance.uc=self.request.user
        return super().form_valid(form)
class precisionlist(LoginRequiredMixin,sin_privilegios,generic.ListView):
    permission_required="inv.view_producto"
    model=historiareporte
    template_name="inv/precision_list.html"
    context_object_name="obj"
    login_url="bases:login"



class reporte_excel(View):
    def get(self,request,*arg,**kwargs):
        productos=Producto.objects.all()
        wb=Workbook()
        ws = wb.active
        ws['B1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['F3'].alignment=Alignment(horizontal="center",vertical="center")
       
        ws['B1'].border=Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B1'].fill=PatternFill(start_color="B8CCE4",end_color="B8CCE4",fill_type="solid")
        ws['B1'].font=Font(name='calibri',bold=True)
        ws['B3'].font=Font(bold=True)
        ws['C3'].font=Font(bold=True)
        ws['D3'].font=Font(bold=True)
        ws['E3'].font=Font(bold=True)
        ws['F3'].font=Font(bold=True)
       
        ws['B1']='REPORTE DE PRODUCTOS'
        ws.merge_cells('B1:F2')
        
        ws.column_dimensions['B'].width=15
        ws.column_dimensions['C'].width=50
        ws.column_dimensions['D'].width=15
        ws.column_dimensions['E'].width=15
        ws.column_dimensions['F'].width=15
       

        ws['B3']='código'
        ws['C3']='descripción'
        ws['D3']='precio'
        ws['E3']='existencia'
        ws['F3']='marca'
       
        
       
        cont=4
        for producto in productos:
            ws.cell(row=cont,column=2).value=producto.codigo
            ws.cell(row=cont,column=3).value=producto.descripcion
            ws.cell(row=cont,column=4).value=producto.precio
            ws.cell(row=cont,column=5).value=producto.existencia
            ws.cell(row=cont,column=6).value=producto.marca.descripcion
           
            
            cont+=1
        nombre_archivo="ReporteProductosExcel.xlsx"
        response= HttpResponse(content_type="application/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response






